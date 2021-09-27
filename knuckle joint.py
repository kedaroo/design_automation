# # FORK PARAMETERS:
# # Front view:
# 1. Enlarged diameter of fork end (1.2d)
# 2. Length of enlarged end (4.5d)
# 3. Outer diameter of eye [d2] (2d)
# 4. Thickness of fork [t1] (0.75d)
# 5. Thickness of rod [t] (1.25d)
# Top view:
# 6. Enlarged diameter of fork end (1.2d)
# 7. Inner diameter of eye [d1] (d)

# ROD PARAMETERS:
# Top view:
# 1. Inner diameter of eye [d1] (d)
# 2. Outer diameter of eye [d2] (2d)
# 3. Length of enlarged end (4d)
# 4. Enlarged diameter of rod (1.1d)
# Front view:
# 5. Enlarged diameter of rod (1.1d)
# 6. Thickness of rod [t] (1.25d)

# PIN PARAMETERS:
# Front view:
# 1. Pin head diameter [d3] (1.5d)
# 2. Thickness of head [t2] (0.5d)
# 3. Diameter of pin [d1] (d)
# 4. Length of pin (4d)

# PIN COLLAR PARAMETERS:
# Top view:
# 1. Inner diameter (d)
# 2. Outer diameter [d3] (1.5d)
# 3. Thickness [t2] (0.5d)

import math
pi = math.pi
from openpyxl import Workbook

# Input Parameters
print('==============================================')
print('======DESIGN AUTOMATION OF KNUCKLE JOINT======')  # USE WHAT U USED IN DAD JOKES IN PYTHON PRACTICE TO DISPLAY THIS MSG
print('==============================================')
P = float(input('Enter load in Newtons: ')) # Load
tensile_stress = float(input('Enter permissible tensile stress in MPa: '))
crushing_stress = float(input('Enter permissible crushing stress in MPa: '))
shear_stress = float(input('Enter permissible shear stress in MPa: '))

# sigma_t_fork = list()
# sigma_c_fork = list()
# tau_fork = list()

# sigma_t_rod = list()
# sigma_c_rod = list()
# tau_rod = list()

# p = [i for i in range(100, 5000000, 10000)]

# for i in p:

# P = i # in newtons
# P = 150000
# tensile_stress = 75 # in MPa
# crushing_stress = 150
# shear_stress = 60

# in mm
def diameter_of_rod(P, tensile_stress):
	 d = math.ceil(math.sqrt((P * 4) / (pi * tensile_stress)))
	 if d % 2 == 0:
	 	return d
	 else:
	 	return d + 1

# def diameter_of_pin():


d = diameter_of_rod(P, tensile_stress)
# print(f'Rod Diameter: {d}')

fork_parameters = ['Enlarged diameter of fork end (1.2d)', 'Length of enlarged end (4.5d)',
	'Outer diameter of eye [d2] (2d)', 'Thickness of fork [t1] (0.75d)', 'Thickness of rod [t] (1.25d)',
	'Enlarged diameter of fork end (1.2d)', 'Inner diameter of eye [d1] (d)']
fork_dimensions = [1.2, 4.5, 2, 0.75, 1.25, 1.2, 1]
fork_design = list()

def calculate_fork_parameters(d):
	for i in range(len(fork_parameters)):
		x = math.ceil(d * fork_dimensions[i])
		fork_design.append(x)

calculate_fork_parameters(d)
dictionary = {fork_parameters[i]: fork_design[i] for i in range(len(fork_parameters))}
# print(dictionary)

rod_parameters = ['Inner diameter of eye [d1] (d)', 'Outer diameter of eye [d2] (2d)', 'Length of enlarged end (4d)',
			'Enlarged diameter of rod (1.1d)', 'Enlarged diameter of rod (1.1d)','Thickness of rod [t] (1.25d)']
rod_dimensions = [1, 2, 4, 1.1, 1.1, 1.25]
rod_design = list()
def calculate_rod_parameters(d):
	for i in range(len(rod_parameters)):
		x = math.ceil(d * rod_dimensions[i])
		rod_design.append(x)
	
calculate_rod_parameters(d)
dictionary = {rod_parameters[i]: rod_design[i] for i in range(len(rod_parameters))}
# print(rod_design)

pin_parameters = ['Pin head diameter [d3] (1.5d)', 'Thickness of head [t2] (0.5d)', 
					'Diameter of pin [d1] (d)', 'Length of pin (4d)']
pin_dimensions = [1.5, 0.5, 1, 4]
pin_design = list()
def calculate_pin_parameters(d):
	for i in range(len(pin_parameters)):
		x = math.ceil(d * pin_dimensions[i])
		pin_design.append(x)

calculate_pin_parameters(d)
# dictionary = {rod_parameters[i]: rod_design[i] for i in range(len(rod_parameters))}
# print(pin_design)

pin_collar_parameters = ['Inner diameter (d)', 'Outer diameter [d3] (1.5d)', 'Thickness [t2] (0.5d)']
pin_collar_dimensions = [1, 1.5, 0.5]
pin_collar_design = list()
def calculate_pin_collar_parameters(d):
	for i in range(len(pin_collar_parameters)):
		x = math.ceil(d * pin_collar_dimensions[i])
		pin_collar_design.append(x)

calculate_pin_collar_parameters(d)
# dictionary = {rod_parameters[i]: rod_design[i] for i in range(len(rod_parameters))}
# print(pin_collar_design)



def check_fork_design():
	d1 = fork_design[-1]
	d2 = fork_design[2]
	t1 = fork_design[3]

	sigma_t = P / ((d2 - d1) * (2 * t1))
	sigma_t_fork.append(sigma_t)

	tau = P / ((d2 - d1) * (2 * t1))
	tau_fork.append(tau)

	sigma_c = P / (d1 * (2 * t1) )
	sigma_c_fork.append(sigma_c)

	if sigma_t < tensile_stress and tau < shear_stress and sigma_c < crushing_stress:
		print(f'{sigma_t}: {tensile_stress}')
		print(f'{sigma_c}: {crushing_stress}')
		print(f'{tau}: {shear_stress}')
		print('Fork design is safe')
	else:
		print(f'{sigma_t}: {tensile_stress}')
		print(f'{sigma_c}: {crushing_stress}')
		print(f'{tau}: {shear_stress}')
		print('Fork design is not safe')



def check_rod_design():
	d1 = rod_design[0]
	d2 = rod_design[1]
	t = rod_design[-1]

	sigma_t = P / ((d2 - d1) * t)
	sigma_t_rod.append(sigma_t)

	tau = P / ((d2 - d1) * t)
	tau_rod.append(tau)

	sigma_c = P / (d1 * t) 
	sigma_c_rod.append(sigma_c)

	if sigma_t < tensile_stress and tau < shear_stress and sigma_c < crushing_stress:
		print(f'{sigma_t}: {tensile_stress}')
		print(f'{sigma_c}: {crushing_stress}')
		print(f'{tau}: {shear_stress}')
		print('Fork design is safe')
	else:
		print(f'{sigma_t}: {tensile_stress}')
		print(f'{sigma_c}: {crushing_stress}')
		print(f'{tau}: {shear_stress}')
		print('Fork design is not safe')

# check_rod_design()


def check_pin_design():
	t1 = fork_design[3]
	t = fork_design[4]
	d1 = fork_design[-1]
	sigma_t = ((P / 2) * ((t1 / 3) + (t / 4))) / ((pi / 32) * pow(d1, 3))
	tau = (P * 2) / (pi * pow(d1, 2))
	if sigma_t < tensile_stress and tau < shear_stress:
		print(f'{sigma_t}: {tensile_stress}')
		print(f'{tau}: {shear_stress}')
		print('Pin design is safe')
	else:
		print(f'{sigma_t}: {tensile_stress}')
		print(f'{tau}: {shear_stress}')
		print('Pin design is not safe')

# check_pin_design()


# check_fork_design()
# check_rod_design()


wb = Workbook()
ws = wb.active

ws['A1'] = 'Fork Parameters'
ws['B1'] = 'Dimension in mm'

for i in range(len(fork_parameters)):
	cell1 = 'A' + str(i + 2)
	cell2 = 'B' + str(i + 2)
	ws[cell1] = fork_parameters[i]
	ws[cell2] = fork_design[i]	


ws['A10'] = 'Rod Parameters'
ws['B10'] = 'Dimension in mm'

for i in range(len(rod_parameters)):
	cell1 = 'A' + str(i + 4 + len(fork_parameters))
	cell2 = 'B' + str(i + 4 + len(fork_parameters))
	ws[cell1] = rod_parameters[i]
	ws[cell2] = rod_design[i]	


ws['A18'] = 'Pin Parameters'
ws['B18'] = 'Dimension in mm'

for i in range(len(pin_parameters)):
	cell1 = 'A' + str(i + 6 + len(fork_parameters) + len(rod_parameters))
	cell2 = 'B' + str(i + 6 + len(fork_parameters) + len(rod_parameters))
	ws[cell1] = pin_parameters[i]
	ws[cell2] = pin_design[i]	

ws['A24'] = 'Pin Collar Parameters'
ws['B24'] = 'Dimension in mm'

for i in range(len(pin_collar_parameters)):
	cell1 = 'A' + str(i + 8 + len(fork_parameters) + len(rod_parameters) + len(pin_parameters))
	cell2 = 'B' + str(i + 8 + len(fork_parameters) + len(rod_parameters) + len(pin_parameters))
	ws[cell1] = pin_parameters[i]
	ws[cell2] = pin_design[i]	

# for i in range(28, 31):
# 	cell1 = 'A' + str(i)
# 	cell2 = 'B' + str(i)
# 	ws[cell1] = 'D' + str(i-27)
# 	ws[cell2] = pin_collar_parameters[i-28]

wb.save('Knuckle Joint Parameters.xlsx')

print('Dimensions calculated successfully and saved to the Excel File: Knuckle Joint Parameters.xlsx.')

# import matplotlib.pyplot as plt 

# fig, axes = plt.subplots(1, 2)
# ax1, ax2 = axes.flatten()

# ax1.plot(p, sigma_c_rod)
# ax1.plot(p, sigma_t_rod, color = 'brown')
# ax1.plot(p, tau_rod, color = 'green')

# ax2.plot(p, sigma_c_fork)
# ax2.plot(p, sigma_t_fork, color = 'brown')
# ax2.plot(p, tau_fork, color = 'green')

# plt.show()